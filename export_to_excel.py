'''

    將資料整理並匯出 excel 檔案

'''

from file_function import read_datas
import openpyxl

def export_balance_to_excel() -> None:
    ''' 將資產負債表資料整理成 excel 型式 並匯出 '''
    
    number_list = read_datas('number', 'stock_number.txt')  # 讀取公司股票代號


    for number in number_list[0].split('\n'):
        workbook = openpyxl.Workbook()  # 建立活頁簿
        workbook.create_sheet('第二季') # 建立第二個工作表

        sheet1 = workbook.worksheets[0] # 取得第一個工作表
        sheet1.title = '第一季'     # 更改工作表名稱

        sheet2 = workbook.worksheets[1] # 取得第二個工作表    

        # 取得資產負債表資訊
        datas = read_datas(str(number), 'balance_sheet.json')
        
        s1_data_list = []   # 儲存 第一季 資料 list
        s2_data_list = []   # 儲存 第二季 資料 list

        for data in datas:
            items = list(data.items())[0]   # 將字典轉為 list

            # 將 str 型態轉為 float (123,456,1.12 -> 1234561.12)
            try:
                s1_values = float(items[1][0].replace(',',''))

             # 如果欄位沒有資料 (-) 則設為 0
            except:
                s1_values = 0

            try:
                s1_percent = float(items[1][1].replace(',',''))
            
            except:
                s1_percent = 0

            try:
                s2_values = float(items[1][2].replace(',', ''))
            
            except:
                s2_values = 0
            
            try:
                s2_percent = float(items[1][3].replace(',', ''))

            except:
                s2_percent = 0

            
            s1_values_list = [
                items[0],   # 欄位名稱
                s1_values,  # 第一季資料
            ]

            s1_percent_list = [
                '%',        # 自己加上百分比欄位
                s1_percent  # 第一季百分比資料
            ]

            s2_values_list = [
                items[0],   # 欄位名稱
                s2_values   # 第二季資料
            ]

            s2_percent_list = [
                '%',        # 自己加上百分比欄位
                s2_percent  # 第二季百分比資料
            ]
            
            # 添加第一季資料，以便後續 excel 處理
            s1_data_list.append(s1_values_list)
            s1_data_list.append(s1_percent_list)

            # 添加第二季資料，以便後續 excel 處理
            s2_data_list.append(s2_values_list)
            s2_data_list.append(s2_percent_list)
        
        # 將資料寫入 excel
        for i in s1_data_list:
            sheet1.append(i)

        for j in s2_data_list:
            sheet2.append(j)

        # 匯出 excel 檔案
        workbook.save(f'DATA/{number}/{number}_balance.xlsx')
    
    return None


def export_income_to_excel(seasion : str) -> None:
    ''' 將 損益表 資料整理成 excel 型式 並匯出 (seasion = 20211 or 20212) '''

    names = read_datas('name', 'stock_name.txt')    # 讀取 公司名稱 檔案
    numbers = read_datas('number', 'stock_number.txt')  # 讀取 公司股票代號 檔案

    # 將讀取的檔案轉為 list 格式
    name_list = [name.split('\n') for name in names][0] 
    number_list = [number.split('\n') for number in numbers][0]


    company_and_number = [' ']  # 存放 公司名稱和股票代碼的 list
    
    # 將公司名稱和股票代碼合併 (台積電 (2330) )
    for index, data in enumerate(number_list):
        company_and_number.append(f'{name_list[index]}  ({data})')

    
    # 定義要匯出的資料 list
    revenue_list = ['營業收入']
    GPM_list = ['營業毛利(毛損)']
    NPFTP_list = ['本期淨利(淨損)']
    BEPS_list = ['基本每股盈餘']

    # 依序讀取股票代碼
    for number in number_list:

        # 建立 excel 的相關物件
        workbook = openpyxl.Workbook()
        sheet = workbook.worksheets[0]
        sheet.title = seasion

        # 讀取公司的損益表 
        datas = read_datas(str(number), f'income_statement_{seasion}.json')
        
        # 如果公司沒有損益表，則將資料設為 0，並跳到下一次迭代
        if datas == []:
            revenue_list.append(0)
            GPM_list.append(0)
            NPFTP_list.append(0)
            BEPS_list.append(0)
            continue


        title_list = [] # 存放 損益表 的欄位名稱 list

        # 依序獲得欄位名稱
        for title in datas:
            title_list.append(list(title.keys())[0])

        revenue_index = title_list.index('營業收入') # 營業收入的 index 值
        GPM_index = title_list.index('營業毛利(毛損)')   # 營業毛利的 index 值
        NPFTP_index = title_list.index('本期淨利(淨損)') # 本期淨利的 index 值
        BEPS_index = title_list.index('基本每股盈餘')    # 基本每股盈餘的 index 值


        # 取得欄位的資料，並將 str 轉為 float (123,258,1.23 -> 1232581.23)
        try:
            
            revenue = float(list(datas[revenue_index].values())[0].replace(',', ''))

        # 如果欄位沒有資料 (-) 則設為 0
        except:
            revenue = 0
        
        try:
            GPM = float(list(datas[GPM_index].values())[0].replace(',', ''))
        
        except:
            GPM = 0
        
        try:
            NPFTP = float(list(datas[NPFTP_index].values())[0].replace(',', ''))

        except:
            NPFTP = 0

        try:
            BEPS = float(list(datas[BEPS_index].values())[0].replace(',', ''))

        except:
            BEPS = 0
        

        # 儲存到相關 list 
        revenue_list.append(float(revenue))
        GPM_list.append(GPM)
        NPFTP_list.append(NPFTP)
        BEPS_list.append(BEPS)

    # 將資料寫入到 excel 中
    sheet.append(company_and_number)
    sheet.append(revenue_list)
    sheet.append(GPM_list)
    sheet.append(NPFTP_list)
    sheet.append(BEPS_list)

    # 匯出 excel 文件
    workbook.save(f'DATA/income/{seasion}.xlsx')

    return None
